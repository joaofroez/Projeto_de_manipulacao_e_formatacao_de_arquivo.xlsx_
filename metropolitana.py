import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

def ajustar_coluna_local(df):
    if 'Local' in df.columns:
        df[['Parte1', 'Parte2', 'Parte3','Parte4']] = df['Local'].str.split('/', expand=True, n=3)
   
        df = df.drop(columns=['Parte1', 'Parte2', 'Local','Parte4'], inplace=True)
    return df

def renomear_col(df):
    df.rename(columns={'Parte3':'Local'}, inplace=True)

def abreviar_data(df):
    df['Criada em'] = df['Criada em'].dt.date

def add_gerencias(df, df2):
    df = pd.merge(df, df2, on='Local', how='left')
    return df

def filtrar_metropolitana(df):
   df_metropolitana = df[df['GERÊNCIA'].isna()]
   return df_metropolitana 

def ajustar_posicao_col(df):
    drm = df.columns[-1]
    df_metropolitana_ordenada = list(df.columns[:-1])
    df_metropolitana_ordenada.insert(4, drm)
    df = df[df_metropolitana_ordenada]
    return df
def alterar_erros(df):
    df['GERÊNCIA'] = df['GERÊNCIA'].fillna('DRM')

def add_nova_linha(df):
    total_linhas = len(df)
    cabecalho_linha = pd.DataFrame([df.columns], columns=df.columns)
    linha_vazia = pd.DataFrame([[''] * len(df.columns)], columns=df.columns)   
    nova_linha = pd.DataFrame([['Solicitações Comerciais Área Metropolitana', '', '', '', 'Total:', total_linhas, '', '', '', '', '']],
                          columns=df.columns)
    df = pd.concat([nova_linha, linha_vazia, cabecalho_linha, df_metropolitana], ignore_index=True)
    df.columns = [None] * df.shape[1]
    return df

def criar_tabela(df):
    num_linhas = len(df) + 1
    num_colunas = len(df.columns)
    ultimo_nome_coluna = chr(64 + num_colunas) 
    ref = f"A4:{ultimo_nome_coluna}{num_linhas}"
    tabela = Table(displayName="Tabela1", ref=ref)
    estilo = TableStyleInfo(
    name="TableStyleMedium9",  
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
    )
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

def add_img():
    img = Image('cedae_img.jpg')
    ws.add_image(img, 'A1')

def tirar_bordas_e_grades():    
    intervalo = 'A1:K1'
    bordas = Border(left=Side(style='none'), 
                     right=Side(style='none'), 
                     top=Side(style='none'), 
                     bottom=Side(style='none'))
    for linha in ws[intervalo]:
        for cell in linha:
            cell.border = bordas

def aumentar_dimensao_col_linha():
   for linha in ws.iter_rows(min_row=1, max_row=2):  # Aumenta a altura das linhas 1 a 2
    ws.row_dimensions[linha[0].row].height = 60  
    
    for col in ws.columns:
        largura_max = 0
        col = [cell for cell in col]  # Lista para fácil manipulação
        for cell in col:
            try:
                if len(str(cell.value)) > largura_max:
                    largura_max = len(str(cell.value))  # Calcula o comprimento máximo
            except:
                pass
        largura_ajus = (largura_max + 3)  # Adiciona um pequeno espaço
        ws.column_dimensions[col[0].column_letter].width = largura_ajus
def alterar_format_cel(*cells):
    estilo_fonte = Font(size=24, color='FFFFFF')
    estilo_fundo = PatternFill(start_color='5E90D2', end_color='5E90D2', fill_type='solid')
    
    for cell in cells:
        cell.font = estilo_fonte
        cell.fill = estilo_fundo

def juntar_cells(cell1,cell2):
    conteudo_junto = f"{ws[cell1].value} {ws[cell2].value}"
    ws[cell1] = conteudo_junto
    ws[cell2].value = None
def alinhar_e_centralizar():
    centralizado = Alignment(horizontal='center', vertical='center')
    max_linhas = ws.max_row
    max_col = ws.max_column
    for linha in ws.iter_rows(min_row=1, max_row=max_linhas, min_col=1, max_col=max_col):
        for cell in linha:
            cell.alignment = centralizado

#------------------------------CARREGAMENTO DAS PLANILHAS------------------------------------------
caminho_arquivo = r'sn_customerservice_task_(20).xlsx'
caminho_arquivo_procx = r'Gerencias_panda.xlsx'

df = pd.read_excel(caminho_arquivo, engine='openpyxl')
df_gerencias = pd.read_excel(caminho_arquivo_procx, engine='openpyxl')

#------------------------------MANIPULAÇÃO DOS DADOS------------------------------------------
ajustar_coluna_local(df)
renomear_col(df)
abreviar_data(df)
df = add_gerencias(df, df_gerencias)
df_metropolitana = filtrar_metropolitana(df)
df_metropolitana = ajustar_posicao_col(df_metropolitana)
alterar_erros(df_metropolitana)
df_metropolitana = add_nova_linha(df_metropolitana)

#------------------------------FORMATAÇÃO DA PLANILHA-----------------------------------------
df_metropolitana.to_excel('Solicitações Comerciais Área Metropolitana.xlsx', index=False)
wb = load_workbook('Solicitações Comerciais Área Metropolitana.xlsx')
ws = wb.active
criar_tabela(df_metropolitana)
add_img()
tirar_bordas_e_grades()
aumentar_dimensao_col_linha()
alterar_format_cel(ws['A2'], ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'])
juntar_cells('E2', 'F2')
ws.merge_cells('A2:C2')
ws.merge_cells('E2:F2')
alinhar_e_centralizar()
ws.sheet_view.showGridLines = False

#------------------------------SALVAMENTO DA PLANILHA-----------------------------------------
wb.save('Solicitações Comerciais Área Metropolitana.xlsx')
